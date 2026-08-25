"""「フラット行エクスポート」形式 (A列=カテゴリ/見出し、B列以降=値) の Excel を読み取り、
連続する同一Aラベルの行を1ブロックにまとめる。

対象フォーマットは今回の `スマスロ ミリオンゴッド-神々の軌跡-_解析.xlsx` のような、
Webページをそのまま1行1事実で書き出したスプレッドシート。将来別サイトの
データでも同様の「ラベル+値の羅列」であれば流用できるよう、この機種固有の
語彙には一切依存しない汎用実装にしている。

このモジュールは読み取り専用。元ファイルは一切変更しない。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from pachislot_ai.ingestion.value_parsing import looks_like_sentence

_PAGE_MARKER_PREFIX = "--- PAGE"
_HEADER_WORD_MAX_LEN = 12


@dataclass(frozen=True, slots=True)
class RawRow:
    row_idx: int
    a: object
    b: object
    c: object
    d: object
    e: object
    f: object
    g: object

    def non_key_cells(self) -> list[object]:
        return [self.b, self.c, self.d, self.e, self.f, self.g]


@dataclass(slots=True)
class RawBlock:
    label: str
    rows: list[RawRow] = field(default_factory=list)
    section_title: str | None = None
    page: str | None = None

    @property
    def row_ref(self) -> str:
        first = self.rows[0].row_idx
        last = self.rows[-1].row_idx
        return str(first) if first == last else f"{first}-{last}"


def read_blocks(path: Path, *, sheet_name: str | None = None) -> list[RawBlock]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = _read_rows(ws)
    finally:
        wb.close()
    blocks = _group_into_blocks(rows)
    return _merge_shape3_tables(blocks)


def _read_rows(ws) -> list[RawRow]:  # noqa: ANN001
    rows: list[RawRow] = []
    for row in ws.iter_rows():
        values = [c.value for c in row]
        while len(values) < 7:
            values.append(None)
        if all(v is None for v in values[:7]):
            continue
        row_idx = row[0].row
        rows.append(RawRow(row_idx, *values[:7]))
    return rows


def _group_into_blocks(rows: list[RawRow]) -> list[RawBlock]:
    blocks: list[RawBlock] = []
    current_label: str | None = None
    current_rows: list[RawRow] = []
    current_page: str | None = None
    current_section: str | None = None

    def flush() -> None:
        nonlocal current_rows, current_label
        if current_rows and current_label is not None:
            blocks.append(
                RawBlock(
                    label=current_label,
                    rows=list(current_rows),
                    section_title=current_section,
                    page=current_page,
                )
            )
        current_rows = []
        current_label = None

    for r in rows:
        a = r.a
        if isinstance(a, str) and a.strip().startswith(_PAGE_MARKER_PREFIX):
            flush()
            current_page = a.strip()
            continue

        if a is None or (isinstance(a, str) and a.strip() == ""):
            # A列が空の行は想定外だが、直前ブロックへの継続として扱う（データを取りこぼさない）
            if current_rows:
                current_rows.append(r)
            continue

        label = str(a).strip()

        if all(v is None for v in r.non_key_cells()):
            # A列のみ埋まっている見出し単独行 -> ブロック化せずセクション見出しとして記録
            flush()
            current_section = label
            continue

        if label == current_label:
            current_rows.append(r)
        else:
            flush()
            current_label = label
            current_section_for_block = current_section
            current_rows = [r]
            current_section = current_section_for_block

    flush()
    return blocks


def _looks_like_header_word(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or looks_like_sentence(text):
        return False
    return len(text) <= _HEADER_WORD_MAX_LEN


def _prepend(key_value: object, row: RawRow) -> RawRow:
    """A列自体がデータ値である行を、通常のB列開始レイアウトへずらして変換する。"""
    shifted = [key_value, *row.non_key_cells()[:-1]]
    return RawRow(row.row_idx, None, *shifted)


def _merge_shape3_tables(blocks: list[RawBlock]) -> list[RawBlock]:
    """A列自体が行キー (例: パターン名) になっている表形式を検出し、1ブロックへ再構成する。

    大半のブロックは A列に「テーブル名の繰り返し」が入り、実データは B列以降に
    並ぶ (Shape1/2)。一部の演出法則テーブルだけは A列自体が「パターン」等の
    実データ列として使われており (Shape3)、そのままでは1行=1ブロックに
    分解されてしまう。ヘッダー行候補 (A列・B列以降がすべて短いラベル語) を
    見つけたら、後続ブロックが同じ列数で続く限り1つの表として束ね直す。
    """
    merged: list[RawBlock] = []
    i = 0
    while i < len(blocks):
        block = blocks[i]
        reconstructed = _try_reconstruct_shape3(blocks, i)
        if reconstructed is not None:
            new_block, next_i = reconstructed
            merged.append(new_block)
            i = next_i
            continue
        merged.append(block)
        i += 1
    return merged


def _try_reconstruct_shape3(
    blocks: list[RawBlock], index: int
) -> tuple[RawBlock, int] | None:
    header_block = blocks[index]
    if len(header_block.rows) != 1:
        return None
    header_row = header_block.rows[0]
    header_extra = [c for c in header_row.non_key_cells() if c is not None]
    if not header_extra:
        return None
    if not _looks_like_header_word(header_block.label):
        return None
    if not all(_looks_like_header_word(c) for c in header_extra):
        return None

    col_count = 1 + len(header_extra)

    data_blocks: list[RawBlock] = []
    j = index + 1
    while j < len(blocks):
        candidate = blocks[j]
        if candidate.page != header_block.page:
            break
        if candidate.section_title != header_block.section_title:
            # 見出し (セクション) をまたいだ先に同じ列形状のテーブルが再度現れても、
            # 別の表として扱う (例: 表モード別GG当選率が複数モードに分かれて連続する場合)
            break
        if not _looks_like_header_word(candidate.label):
            break
        ok = True
        for r in candidate.rows:
            filled = [c for c in r.non_key_cells() if c is not None]
            if len(filled) + 1 != col_count:
                ok = False
                break
            # 値が長文・説明文らしい場合は表の継続データ行とはみなさない
            # (短いラベルの繰り返しに見えても、実際は別の解説ブロックであるケースを除外する)
            if any(isinstance(c, str) and looks_like_sentence(c) for c in filled):
                ok = False
                break
        if not ok:
            break
        data_blocks.append(candidate)
        j += 1

    if len(data_blocks) < 2:
        return None

    new_rows = [_prepend(header_block.label, header_row)]
    for data_block in data_blocks:
        for r in data_block.rows:
            new_rows.append(_prepend(data_block.label, r))

    label = header_block.section_title or header_block.label
    new_block = RawBlock(
        label=label, rows=new_rows, section_title=header_block.section_title, page=header_block.page
    )
    return new_block, j
